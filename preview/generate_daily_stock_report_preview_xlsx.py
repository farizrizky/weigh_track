from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


OUTPUT_PATH = Path(__file__).with_name("daily_stock_report_preview.xlsx")

GREEN_HEADER = "C8DCA8"
GREEN_SECTION = "DBE7C5"
GREEN_SUBSECTION = "EEF4E5"
ORANGE_TRANSFER = "FFF1D6"
DARK_TEXT = "172033"
MUTED_TEXT = "4B5563"
TRANSFER_TEXT = "A44200"
PERCENT_TEXT = "C5221F"
WHITE = "FFFFFF"

THIN_SIDE = Side(style="thin", color="4B5563")
TABLE_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def data_row(label, values, unit="Kg Basah", kind="data"):
    return {
        "kind": kind,
        "values": ["", label, unit, *values],
    }


rows = [
    {"kind": "section", "no": "I", "label": "PEKERJAAN YANG DILAKSANAKAN"},
    data_row("Stock Awal Produksi", [1000, 800, 600, 700, 700, 1800, 2000, 3800]),
    data_row(
        "Stock Awal Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 200],
    ),
    data_row("Jumlah Stock Awal", [1000, 800, 600, 700, 700, 1800, 2000, 4000], kind="total"),
    {"kind": "subsection", "label": "Penimbangan Kebun"},
    data_row("Hari ini", [100, 80, 120, 50, 60, 180, 230, 410]),
    data_row("Sampai dengan hari ini", [520, 430, 610, 300, 360, 950, 1270, 2220], kind="total"),
    {"kind": "subsection", "label": "Penimbangan Gudang"},
    data_row("Gudang Induk", [98, 80, 118, "-", 60, 178, 178, 356]),
    data_row("Gudang Transit", ["-", "-", "-", 50, "-", "-", 50, 50]),
    data_row("Jumlah Penerimaan Gudang", [98, 80, 118, 50, 60, 178, 228, 406], kind="total"),
    data_row("Sampai dengan hari ini", [510, 428, 600, 295, 355, 938, 1250, 2188], kind="total"),
    {"kind": "subsection", "label": "Pengeluaran Produksi"},
    data_row("Pengiriman dari Lot Produksi", [300, 220, 180, 48, 120, 520, 348, 868]),
    data_row(
        "Pengiriman dari Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 98],
    ),
    data_row(
        "Susut Transfer Antar Gudang - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 2],
        kind="transfer",
    ),
    data_row("Jumlah Pengeluaran Hari ini", [300, 220, 180, 48, 120, 520, 348, 968], kind="total"),
    {"kind": "section", "no": "II", "label": "SUSUT PRODUKSI"},
    {"kind": "subsection", "label": "Susut Timbang dari Kebun ke Gudang"},
    data_row("Hari ini", [2, 0, 2, 0, 0, 2, 2, 4]),
    data_row("Hari ini", [0.02, 0.0, 0.0167, 0.0, 0.0, 0.0111, 0.0087, 0.0098], unit="%", kind="percentage"),
    data_row("Sampai dengan hari ini", [10, 2, 10, 5, 5, 12, 20, 32], kind="total"),
    {"kind": "subsection", "label": "Susut dari Gudang ke Pengiriman"},
    data_row("Susut Penyimpanan Hari ini", [5, 3, 2, 4, 1, 8, 7, 15]),
    data_row(
        "Susut Transfer - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 2],
        kind="transfer",
    ),
    data_row("Jumlah Susut Hari ini", [5, 3, 2, 4, 1, 8, 7, 17], kind="total"),
    data_row("Jumlah Susut Hari ini", [0.0046, 0.0034, 0.0028, 0.0042, 0.0013, 0.0040, 0.0037, 0.0039], unit="%", kind="percentage"),
    data_row("Susut Penyimpanan s.d. Hari ini", [12, 8, 12, 15, 7, 20, 34, 54]),
    data_row(
        "Susut Transfer s.d. Hari ini - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 6],
        kind="transfer",
    ),
    data_row("Jumlah Susut s.d. Hari ini", [12, 8, 12, 15, 7, 20, 34, 60], kind="total"),
    {"kind": "section", "no": "III", "label": "STOCK PRODUKSI"},
    data_row("Saldo Awal Produksi", [1000, 800, 600, 700, 700, 1800, 2000, 3800]),
    data_row(
        "Saldo Awal Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 200],
    ),
    data_row("Jumlah Saldo Awal", [1000, 800, 600, 700, 700, 1800, 2000, 4000], kind="total"),
    data_row("Produksi Masuk", [98, 80, 118, 50, 60, 178, 228, 406]),
    data_row("Pengiriman dari Lot Produksi", [300, 220, 180, 48, 120, 520, 348, 868]),
    data_row(
        "Pengiriman dari Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 98],
    ),
    data_row("Susut Penyimpanan", [5, 3, 2, 4, 1, 8, 7, 15]),
    data_row(
        "Mutasi Keluar ke Lot Transit",
        ["-", "-", "-", -100, "-", "-", -100, -100],
        kind="transfer",
    ),
    data_row(
        "Mutasi Masuk Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 100],
        kind="transfer",
    ),
    data_row(
        "Susut Transfer - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 2],
        kind="transfer",
    ),
    data_row("Saldo Akhir Produksi", [793, 657, 536, 598, 639, 1450, 1773, 3223]),
    data_row(
        "Saldo Akhir Lot Transit - Gudang Induk Gembung",
        ["-", "-", "-", "-", "-", "-", "-", 200],
    ),
    data_row("Jumlah Saldo Akhir", [793, 657, 536, 598, 639, 1450, 1773, 3423], kind="total"),
    {"kind": "section", "no": "IV", "label": "RINGKASAN STOCK PER LOKASI"},
    data_row("Gudang Induk Sebayur", [793, 657, "-", "-", "-", 1450, "-", 1450]),
    data_row("Gudang Induk Gembung - Lot Produksi", ["-", "-", 536, 598, 639, "-", 1773, 1773]),
    data_row("Gudang Induk Gembung - Lot Transit", ["-", "-", "-", "-", "-", "-", "-", 200]),
    data_row("Jumlah", [793, 657, 536, 598, 639, 1450, 1773, 3423], kind="total"),
]


workbook = Workbook()
sheet = workbook.active
sheet.title = "Laporan Harian"
sheet.sheet_view.showGridLines = False
sheet.freeze_panes = "D6"

sheet.merge_cells("A1:C2")
sheet["A1"] = "JULANG\nPLANTATIONS"
sheet["A1"].font = Font(name="Arial", size=18, bold=True, color="247A3B")
sheet["A1"].alignment = Alignment(vertical="center", wrap_text=True)

sheet.merge_cells("G1:K1")
sheet["G1"] = "LAPORAN HARIAN OPERASIONAL DAN STOCK"
sheet["G1"].font = Font(name="Arial", size=15, bold=True, color=DARK_TEXT)
sheet["G1"].alignment = Alignment(horizontal="right", vertical="center")

sheet.merge_cells("G2:K2")
sheet["G2"] = "Tanggal: 02/08/2026   |   Perusahaan: PT. Julang Plantations"
sheet["G2"].font = Font(name="Arial", size=8, color=MUTED_TEXT)
sheet["G2"].alignment = Alignment(horizontal="right", vertical="center")

sheet.row_dimensions[1].height = 22
sheet.row_dimensions[2].height = 18
sheet.row_dimensions[3].height = 6

sheet.merge_cells("A4:A5")
sheet.merge_cells("B4:B5")
sheet.merge_cells("C4:C5")
sheet.merge_cells("D4:H4")
sheet.merge_cells("I4:J4")
sheet.merge_cells("K4:K5")

headers = {
    "A4": "No",
    "B4": "Uraian",
    "C4": "Satuan",
    "D4": "Sub Divisi",
    "I4": "Divisi",
    "K4": "Total",
    "D5": "I",
    "E5": "III",
    "F5": "IV",
    "G5": "V",
    "H5": "VI",
    "I5": "Sebayur",
    "J5": "Gembung",
}
for coordinate, value in headers.items():
    cell = sheet[coordinate]
    cell.value = value
    cell.font = Font(name="Arial", size=8, bold=True, color=DARK_TEXT)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in sheet.iter_rows(min_row=4, max_row=5, min_col=1, max_col=11):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=GREEN_HEADER)
        cell.border = TABLE_BORDER

current_row = 6
for row_data in rows:
    kind = row_data["kind"]

    if kind == "section":
        sheet.cell(current_row, 1, row_data["no"])
        sheet.cell(current_row, 2, row_data["label"])
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
    elif kind == "subsection":
        sheet.cell(current_row, 2, row_data["label"])
    else:
        for column, value in enumerate(row_data["values"], start=1):
            sheet.cell(current_row, column, value)

    for column in range(1, 12):
        cell = sheet.cell(current_row, column)
        cell.border = TABLE_BORDER
        cell.font = Font(name="Arial", size=8, color=DARK_TEXT)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if column == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif column >= 4:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;[Red]-#,##0;0"

    if kind == "section":
        fill = PatternFill("solid", fgColor=GREEN_SECTION)
        for column in range(1, 12):
            sheet.cell(current_row, column).fill = fill
            sheet.cell(current_row, column).font = Font(name="Arial", size=8, bold=True, color=DARK_TEXT)
    elif kind == "subsection":
        fill = PatternFill("solid", fgColor=GREEN_SUBSECTION)
        for column in range(1, 12):
            sheet.cell(current_row, column).fill = fill
            sheet.cell(current_row, column).font = Font(name="Arial", size=8, bold=True, color=DARK_TEXT)
    elif kind == "total":
        fill = PatternFill("solid", fgColor=GREEN_SECTION)
        for column in range(1, 12):
            sheet.cell(current_row, column).fill = fill
            sheet.cell(current_row, column).font = Font(name="Arial", size=8, bold=True, color=DARK_TEXT)
    elif kind == "transfer":
        fill = PatternFill("solid", fgColor=ORANGE_TRANSFER)
        for column in range(1, 12):
            sheet.cell(current_row, column).fill = fill
            sheet.cell(current_row, column).font = Font(name="Arial", size=8, bold=True, color=TRANSFER_TEXT)
    elif kind == "percentage":
        for column in range(1, 12):
            sheet.cell(current_row, column).font = Font(name="Arial", size=8, color=PERCENT_TEXT)
        for column in range(4, 12):
            if isinstance(sheet.cell(current_row, column).value, (int, float)):
                sheet.cell(current_row, column).number_format = "0.00%"

    sheet.row_dimensions[current_row].height = 16
    current_row += 1

footer_row = current_row + 1
sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
sheet.cell(
    footer_row,
    1,
    "Saldo akhir = saldo awal + produksi masuk - pengiriman - susut penyimpanan + mutasi internal - susut transfer.",
)
sheet.cell(footer_row, 1).font = Font(name="Arial", size=7, color=MUTED_TEXT)
sheet.cell(footer_row, 1).alignment = Alignment(vertical="center")
sheet.merge_cells(start_row=footer_row, start_column=9, end_row=footer_row, end_column=11)
sheet.cell(footer_row, 9, "Preview data contoh")
sheet.cell(footer_row, 9).font = Font(name="Arial", size=7, color=MUTED_TEXT)
sheet.cell(footer_row, 9).alignment = Alignment(horizontal="right", vertical="center")

column_widths = {
    "A": 5,
    "B": 43,
    "C": 12,
    "D": 11,
    "E": 11,
    "F": 11,
    "G": 11,
    "H": 11,
    "I": 13,
    "J": 14,
    "K": 16,
}
for column, width in column_widths.items():
    sheet.column_dimensions[column].width = width

sheet.auto_filter.ref = f"A4:K{current_row - 1}"
sheet.print_title_rows = "4:5"
sheet.print_area = f"A1:K{footer_row}"
sheet.page_setup.orientation = "landscape"
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToWidth = 1
sheet.page_setup.fitToHeight = 1
sheet.sheet_properties.pageSetUpPr.fitToPage = True
sheet.page_margins = PageMargins(
    left=0.25,
    right=0.25,
    top=0.3,
    bottom=0.3,
    header=0.1,
    footer=0.1,
)

workbook.save(OUTPUT_PATH)
print(OUTPUT_PATH)
