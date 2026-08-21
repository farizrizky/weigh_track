# -*- coding: utf-8 -*-

import base64
import io

from odoo import _, fields, models
from odoo.exceptions import ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class FieldProductionImportWizard(models.TransientModel):
    _name = "wt.field.production.import.wizard"
    _description = "Wizard Import Excel Produksi Field"

    production_id = fields.Many2one(
        "wt.field.production",
        string="Produksi Field",
        required=True,
        ondelete="cascade",
        readonly=True,
    )
    file_data = fields.Binary(
        string="File Excel",
        help="Upload file Excel (.xlsx). Kolom yang diperlukan: Field, Produksi Hari Ini (kg).",
    )
    file_name = fields.Char(string="Nama File")

    def _get_field_code_map(self):
        """
        Buat mapping: kode/nama field → field_line record
        dari field_line_ids produksi yang dipilih.
        """
        mapping = {}
        for line in self.production_id.field_line_ids:
            field_rec = line.field_id
            # Index by field name/code (case-insensitive)
            key = (field_rec.display_name or "").strip().upper()
            if key:
                mapping[key] = line
            # Also index by field code if different
            code_key = (field_rec.name or "").strip().upper()
            if code_key and code_key != key:
                mapping[code_key] = line
        return mapping

    def action_import(self):
        """Parse Excel dan update today_production_weight di baris field produksi."""
        self.ensure_one()

        if openpyxl is None:
            raise ValidationError(
                _("Package openpyxl belum terinstall. "
                  "Jalankan: pip install openpyxl")
            )

        if not self.file_data:
            raise ValidationError(_("Harap upload file Excel terlebih dahulu."))

        # Decode file
        try:
            raw = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError(
                _("File tidak dapat dibaca. Pastikan file berformat .xlsx.\nDetail: %s") % str(exc)
            )

        ws = wb.active

        # Deteksi baris header — cari baris yang mengandung "field" (case-insensitive)
        header_row_index = None
        col_field = None
        col_weight = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
            row_values = [(cell.value or "").strip().upper() if cell.value else "" for cell in row]
            for col_idx, val in enumerate(row_values):
                if "FIELD" == val:
                    col_field = col_idx
                if "PRODUKSI HARI INI" in val or "HARI INI" in val:
                    col_weight = col_idx
            if col_field is not None and col_weight is not None:
                header_row_index = row_idx
                break

        if header_row_index is None:
            raise ValidationError(
                _(
                    "Header kolom tidak ditemukan.\n"
                    "Pastikan file Excel memiliki kolom dengan nama:\n"
                    "  - 'Field'\n"
                    "  - 'Produksi Hari Ini (kg)' atau 'Hari Ini'"
                )
            )

        field_map = self._get_field_code_map()
        updated_count = 0
        not_found = []

        for row in ws.iter_rows(min_row=header_row_index + 1):
            field_cell = row[col_field]
            weight_cell = row[col_weight]

            field_val = (field_cell.value or "")
            if not field_val:
                continue
            field_key = str(field_val).strip().upper()

            # Cari matching line
            line = field_map.get(field_key)
            if not line:
                not_found.append(str(field_val).strip())
                continue

            # Parse berat
            try:
                weight = float(weight_cell.value or 0.0)
            except (TypeError, ValueError):
                weight = 0.0

            line.today_production_weight = weight
            updated_count += 1

        wb.close()

        # Susun pesan hasil
        msg_parts = [_("Import selesai. %d baris berhasil diperbarui.") % updated_count]
        if not_found:
            msg_parts.append(
                _("Field berikut tidak ditemukan di daftar baris produksi:\n%s")
                % "\n".join("  • %s" % f for f in not_found)
            )

        # Tampilkan notifikasi
        message = "\n\n".join(msg_parts)
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Excel"),
                "message": message,
                "type": "success" if not not_found else "warning",
                "sticky": bool(not_found),
                "next": {"type": "ir.actions.act_window_close"},
            },
        }

    def action_download_template(self):
        """Generate dan download template Excel kosong untuk panduan import."""
        self.ensure_one()

        try:
            import xlsxwriter
        except ImportError:
            raise ValidationError(_("Package xlsxwriter belum terinstall."))

        prod = self.production_id
        division_name = prod.division_id.display_name or "-"
        company_name  = prod.company_id.name or ""

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        sheet = workbook.add_worksheet("Produksi Field")

        # --- Formats ---
        title_fmt = workbook.add_format({
            "bold": True, "font_size": 14, "align": "center", "valign": "vcenter",
        })
        info_label_fmt = workbook.add_format({
            "bold": True, "valign": "vcenter", "bg_color": "#F1F5F9",
            "border": 1, "font_size": 10,
        })
        info_value_fmt = workbook.add_format({
            "valign": "vcenter", "border": 1, "font_size": 10,
        })
        header_fmt = workbook.add_format({
            "bold": True, "align": "center", "valign": "vcenter",
            "border": 1, "bg_color": "#E2E8F0", "font_size": 10,
        })
        text_fmt = workbook.add_format({
            "border": 1, "valign": "vcenter", "font_size": 10,
        })
        number_fmt = workbook.add_format({
            "border": 1, "num_format": "#,##0.00", "valign": "vcenter", "font_size": 10,
        })
        note_fmt = workbook.add_format({
            "italic": True, "font_color": "#6B7280", "font_size": 9,
        })

        # --- Column widths ---
        sheet.set_column(0, 0, 16)   # Field
        sheet.set_column(1, 1, 14)   # Clone
        sheet.set_column(2, 2, 10)   # HA
        sheet.set_column(3, 3, 26)   # Produksi Hari Ini

        # Row 0: Judul perusahaan
        sheet.merge_range(0, 0, 0, 3, company_name, title_fmt)
        sheet.set_row(0, 22)

        # Row 1: Judul dokumen
        sheet.merge_range(1, 0, 1, 3, "TEMPLATE INPUT PRODUKSI FIELD", title_fmt)
        sheet.set_row(1, 18)

        # Row 2: blank separator
        sheet.set_row(2, 6)

        # Row 3: Tanggal Produksi — kosong, diisi oleh kepala divisi setiap pakai
        editable_fmt = workbook.add_format({
            "valign": "vcenter", "border": 1, "font_size": 10,
            "bg_color": "#FFFDE7",  # kuning muda → tanda "isi di sini"
            "italic": True, "font_color": "#9E6C00",
        })
        sheet.write(3, 0, "Tanggal Produksi", info_label_fmt)
        sheet.merge_range(3, 1, 3, 3, "[ isi tanggal produksi ]", editable_fmt)
        sheet.set_row(3, 18)

        # Row 4: Divisi — terisi, hanya info referensi (tidak dibaca saat import)
        sheet.write(4, 0, "Divisi", info_label_fmt)
        sheet.merge_range(4, 1, 4, 3, division_name, info_value_fmt)
        sheet.set_row(4, 18)

        # Row 5: blank separator
        sheet.set_row(5, 6)

        # Row 6: Petunjuk
        sheet.merge_range(6, 0, 6, 3,
            "Petunjuk: Isi kolom 'Produksi Hari Ini (kg)'. "
            "JANGAN ubah nama Field, Clone, atau HA.",
            note_fmt,
        )
        sheet.set_row(6, 14)

        # Row 7: Header tabel
        for col, h in enumerate(["Field", "Clone", "HA", "Produksi Hari Ini (kg)"]):
            sheet.write(7, col, h, header_fmt)
        sheet.set_row(7, 18)

        # Freeze panes agar header tidak bergerak saat scroll
        sheet.freeze_panes(8, 0)

        # Rows 8+: Data field lines
        for row_idx, line in enumerate(prod.field_line_ids, start=8):
            sheet.write(row_idx, 0, line.field_id.display_name or "", text_fmt)
            sheet.write(row_idx, 1, line.clone or "", text_fmt)
            sheet.write(row_idx, 2, line.ha or 0.0, number_fmt)
            sheet.write(row_idx, 3, 0.0, number_fmt)

        workbook.close()
        output.seek(0)

        # Nama file: hanya sampai nama divisi (tanpa tanggal — bisa dipakai berulang)
        filename = "Template Produksi Field - %s.xlsx" % division_name
        attachment = self.env["ir.attachment"].create({
            "name": filename,
            "type": "binary",
            "datas": base64.b64encode(output.read()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        })
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "new",
        }
